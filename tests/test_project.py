from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook
import pytest
from scripts.cms_engine import CMSValidationError, load_cms

ROOT=Path(__file__).resolve().parents[1]

def data(): return json.loads((ROOT/"data/content.json").read_text(encoding="utf-8"))

def test_core_catalog_is_complete_and_campaign_free():
    cms=data()
    assert cms["meta"]["version"] == "5.0.0-core"
    assert cms["meta"]["campaigns"] == []
    assert len(cms["contents"]) == len(cms["catalog"]) == 103
    assert sum(item["category"] != "Alimentos" for item in cms["contents"]) == 97
    assert sum(item["category"] == "Alimentos" for item in cms["contents"]) == 6
    assert "uni" + "corn" not in json.dumps(cms,ensure_ascii=False).lower()

def test_every_module_has_an_independent_recipe_and_steps():
    cms=data(); routes={step["route"] for step in cms["steps"]}
    for item in cms["contents"]:
        assert (ROOT/item["productImage"]).is_file()
        assert (ROOT/item["referenceImage"]).is_file()
        assert item["equipment"] and item["rules"]
        assert set(item["routes"].values()) <= routes
    assert any(step["media"].startswith("assets/GIF_CORE/") for step in cms["steps"])

def test_lots_respect_operational_limits():
    expected={"Lote_01_Bebidas":49,"Lote_02_Bebidas":48,"Lote_01_Recetas":49,"Lote_02_Recetas":48,"Lote_01_Alimentos":6,"GIF_CORE":16}
    for folder,count in expected.items():
        files=[p for p in (ROOT/"assets"/folder).iterdir() if p.is_file()]
        assert len(files)==count
        assert len(files)<=100
        assert sum(p.stat().st_size for p in files)<25*1024*1024

def test_cms_is_single_canonical_source():
    cms=ROOT/"outputs/CMS_Guia_Operativa_CORE.xlsx"
    assert cms.is_file()
    assert not (ROOT/"outputs/CMS_Guia_Operativa_v2.xlsx").exists()
    book=load_workbook(cms,read_only=True,data_only=False)
    assert book.sheetnames == ["Inicio","Contenidos","Selectores","Opciones","Rutas","Pasos","Equipo","Normas","Medios","Campanas","Control"]
    assert book["Control"]["B14"].value.startswith("=IF(")
    book.close()

def test_ui_is_core_and_didactic():
    html=(ROOT/"index.html").read_text(encoding="utf-8").lower()
    app=(ROOT/"app.js").read_text(encoding="utf-8").lower()
    styles=(ROOT/"styles.css").read_text(encoding="utf-8").lower()
    combined=html+app+styles
    assert "uni" + "corn" not in combined
    assert "object" + "ives" not in combined
    assert 'id="trainingstage"' in html and 'id="recipegrid"' in html
    assert "reference-dock" in app and "gif_core" in json.dumps(data()).lower()

def test_excel_is_the_exact_source_of_generated_data():
    assert load_cms() == data()

def test_blank_step_media_uses_product_fallback():
    cms=data()
    aerocano=[step for step in cms["steps"] if step["route"]=="aerocano-core"]
    assert aerocano[1]["media"] == ""
    app=(ROOT/"app.js").read_text(encoding="utf-8")
    assert "current.media||state.content.productImage" in app

def test_workflow_checks_before_any_generation():
    workflow=(ROOT/".github/workflows/validate-and-deploy.yml").read_text(encoding="utf-8")
    assert "python scripts/build_content.py\n" not in workflow
    assert "python scripts/build_content.py --check" in workflow

def test_invalid_media_path_is_rejected(tmp_path):
    source=ROOT/"outputs/CMS_Guia_Operativa_CORE.xlsx"
    target=tmp_path/"invalid.xlsx"
    book=load_workbook(source)
    book["Pasos"]["I5"]="assets/GIF_CORE/no-existe.gif"
    book.save(target); book.close()
    with pytest.raises(CMSValidationError, match="no existe"):
        load_cms(target)

def test_navigation_has_keyboard_and_visible_focus_support():
    app=(ROOT/"app.js").read_text(encoding="utf-8")
    styles=(ROOT/"styles.css").read_text(encoding="utf-8")
    assert 'event.key==="ArrowLeft"' in app and 'event.key==="ArrowRight"' in app
    assert 'event.key==="/"' in app
    assert ":focus-visible" in styles
