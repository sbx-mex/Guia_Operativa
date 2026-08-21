from __future__ import annotations

from collections import defaultdict
from pathlib import Path, PurePosixPath
from typing import Any
from itertools import product

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CMS_PATH = ROOT / "outputs" / "CMS_Guia_Operativa_CORE.xlsx"
HEADER_ROW = 4


class CMSValidationError(ValueError):
    """El CMS no se puede publicar de forma segura."""


def _active(value: Any) -> bool:
    return str(value or "SI").strip().upper() in {"SI", "SÍ", "TRUE", "1"}


def _rows(book, sheet: str, required: tuple[str, ...]) -> list[dict[str, Any]]:
    if sheet not in book.sheetnames:
        raise CMSValidationError(f"Falta la pestaña obligatoria: {sheet}")
    ws = book[sheet]
    headers = [str(cell.value or "").strip() for cell in ws[HEADER_ROW]]
    missing = [name for name in required if name not in headers]
    if missing:
        raise CMSValidationError(f"{sheet}: faltan encabezados {', '.join(missing)}")
    result = []
    for values in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result


def _safe_media(path: Any) -> str:
    value = str(path or "").strip().replace("\\", "/")
    pure = PurePosixPath(value)
    if not value or pure.is_absolute() or ".." in pure.parts or not value.startswith("assets/"):
        raise CMSValidationError(f"Ruta de medio insegura: {value or '(vacía)'}")
    return value


def _required_text(row: dict[str, Any], key: str, context: str) -> str:
    value = str(row.get(key) or "").strip()
    if not value:
        raise CMSValidationError(f"{context}: {key} está vacío")
    return value


def _unique(values: list[str], context: str) -> None:
    duplicates = sorted(value for value, count in __import__("collections").Counter(values).items() if count > 1)
    if duplicates:
        raise CMSValidationError(f"{context}: valores duplicados: {', '.join(duplicates[:10])}")


def _existing_media(path: Any, context: str, optional: bool = False) -> str:
    if optional and not str(path or "").strip():
        return ""
    value = _safe_media(path)
    target = ROOT / value
    if not target.is_file():
        raise CMSValidationError(f"{context}: no existe {value}")
    if target.suffix.lower() not in {".webp", ".gif", ".png", ".jpg", ".jpeg", ".mp4", ".pdf"}:
        raise CMSValidationError(f"{context}: extensión no permitida en {value}")
    return value


def load_cms(path: Path = CMS_PATH) -> dict[str, Any]:
    if not path.exists():
        raise CMSValidationError(f"No existe el motor CMS: {path.relative_to(ROOT)}")
    book = load_workbook(path, data_only=True, read_only=True)
    contents_rows = _rows(book, "Contenidos", ("ID_CONTENIDO", "CATEGORIA", "NOMBRE", "IMAGEN_PRODUCTO", "FICHA_REFERENCIA", "ACTIVO"))
    selector_rows = _rows(book, "Selectores", ("ID_CONTENIDO", "ID_SELECTOR", "ETIQUETA", "ORDEN"))
    option_rows = _rows(book, "Opciones", ("ID_CONTENIDO", "ID_SELECTOR", "ID_OPCION", "ETIQUETA", "ORDEN"))
    route_rows = _rows(book, "Rutas", ("ID_CONTENIDO", "CONDICION", "ID_RUTA", "ACTIVO"))
    step_rows = _rows(book, "Pasos", ("ID_RUTA", "ORDEN", "TITULO", "DETALLE", "ICONO", "VALORES", "TIMER_SEG", "ETAPA", "ACTIVO"))
    equipment_rows = _rows(book, "Equipo", ("ID_CONTENIDO", "ORDEN", "ELEMENTO", "ACTIVO"))
    rule_rows = _rows(book, "Normas", ("ID_CONTENIDO", "ORDEN", "NORMA", "ACTIVO"))
    media_rows = _rows(book, "Medios", ("ID_MEDIO", "NOMBRE", "CATEGORIA", "SUBCATEGORIA", "IMAGEN_PRODUCTO", "FICHA_REFERENCIA"))
    campaign_rows = _rows(book, "Campanas", ("ID_CAMPANA", "TITULO", "FECHA_INICIO", "FECHA_FIN", "ZONA_HORARIA", "ID_PRINCIPAL", "ID_SECUNDARIO", "ACTIVO"))
    book.close()

    active_contents = [row for row in contents_rows if _active(row.get("ACTIVO"))]
    content_ids = [str(row["ID_CONTENIDO"]).strip() for row in active_contents]
    _unique(content_ids, "Contenidos/ID_CONTENIDO")
    content_id_set = set(content_ids)

    options: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    labels: dict[str, str] = {}
    for row in option_rows:
        content_id = _required_text(row, "ID_CONTENIDO", "Opciones")
        selector_id = _required_text(row, "ID_SELECTOR", f"Opciones/{content_id}")
        if content_id not in content_id_set:
            raise CMSValidationError(f"Opción enlazada a contenido inactivo o desconocido: {content_id}")
        key = (content_id, selector_id)
        option_id = _required_text(row, "ID_OPCION", f"Opciones/{content_id}/{selector_id}")
        if any(item["id"] == option_id for item in options[key]):
            raise CMSValidationError(f"Opción duplicada: {content_id}/{selector_id}/{option_id}")
        labels[option_id] = str(row["ETIQUETA"]).strip()
        options[key].append({"id": option_id, "order": int(row.get("ORDEN") or 0)})

    selectors: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in selector_rows:
        content_id, selector_id = str(row["ID_CONTENIDO"]).strip(), str(row["ID_SELECTOR"]).strip()
        if content_id not in content_id_set:
            raise CMSValidationError(f"Selector enlazado a contenido inactivo o desconocido: {content_id}")
        choices = [item["id"] for item in sorted(options[(content_id, selector_id)], key=lambda item: item["order"])]
        if not choices:
            raise CMSValidationError(f"Selector sin opciones: {content_id}/{selector_id}")
        selectors[content_id].append({"id": selector_id, "label": str(row["ETIQUETA"]).strip(), "options": choices, "order": int(row.get("ORDEN") or 0)})

    routes: dict[str, dict[str, str]] = defaultdict(dict)
    for row in route_rows:
        if _active(row.get("ACTIVO")):
            content_id = _required_text(row, "ID_CONTENIDO", "Rutas")
            condition = _required_text(row, "CONDICION", f"Rutas/{content_id}")
            route_id = _required_text(row, "ID_RUTA", f"Rutas/{content_id}/{condition}")
            if content_id not in content_id_set:
                raise CMSValidationError(f"Ruta enlazada a contenido inactivo o desconocido: {content_id}")
            if condition in routes[content_id]:
                raise CMSValidationError(f"Condición de ruta duplicada: {content_id}/{condition}")
            routes[content_id][condition] = route_id

    equipment: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for row in equipment_rows:
        if _active(row.get("ACTIVO")):
            equipment[str(row["ID_CONTENIDO"]).strip()].append((int(row.get("ORDEN") or 0), str(row["ELEMENTO"]).strip()))
    rules: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for row in rule_rows:
        if _active(row.get("ACTIVO")):
            rules[str(row["ID_CONTENIDO"]).strip()].append((int(row.get("ORDEN") or 0), str(row["NORMA"]).strip()))

    contents = []
    for row in active_contents:
        content_id = str(row["ID_CONTENIDO"]).strip()
        content_selectors = sorted(selectors[content_id], key=lambda item: item["order"])
        for selector in content_selectors:
            selector.pop("order")
        if not routes[content_id]:
            raise CMSValidationError(f"Contenido sin rutas: {content_id}")
        contents.append({
            "id": content_id, "name": str(row["NOMBRE"]).strip(), "category": str(row["CATEGORIA"]).strip(),
            "subcategory": str(row.get("SUBCATEGORIA") or "General").strip(), "description": str(row.get("DESCRIPCION") or "").strip(),
            "productImage": _existing_media(row["IMAGEN_PRODUCTO"], f"Contenidos/{content_id}/IMAGEN_PRODUCTO"),
            "referenceImage": _existing_media(row["FICHA_REFERENCIA"], f"Contenidos/{content_id}/FICHA_REFERENCIA"),
            "selectors": content_selectors, "routes": routes[content_id],
            "equipment": [value for _, value in sorted(equipment[content_id])],
            "rules": [value for _, value in sorted(rules[content_id])],
        })
        expected_conditions = {"default"} if not content_selectors else {
            "|".join(f"{selector['id']}={choice}" for selector, choice in zip(content_selectors, choices))
            for choices in product(*(selector["options"] for selector in content_selectors))
        }
        actual_conditions = set(routes[content_id])
        if expected_conditions != actual_conditions:
            raise CMSValidationError(
                f"Rutas incompletas en {content_id}: faltan={sorted(expected_conditions-actual_conditions)}, "
                f"sobran={sorted(actual_conditions-expected_conditions)}"
            )

    steps = []
    orders: dict[str, list[int]] = defaultdict(list)
    for row in step_rows:
        if not _active(row.get("ACTIVO")):
            continue
        route = _required_text(row, "ID_RUTA", "Pasos")
        try:
            order = int(row["ORDEN"])
        except (TypeError, ValueError):
            raise CMSValidationError(f"Pasos/{route}: ORDEN debe ser entero") from None
        orders[route].append(order)
        steps.append({"route": route, "order": order, "title": str(row["TITULO"]).strip(),
                      "detail": str(row["DETALLE"]).strip(), "icon": str(row["ICONO"]).strip(),
                      "values": str(row.get("VALORES") or "").strip(), "timer": int(row.get("TIMER_SEG") or 0),
                      "stage": str(row.get("ETAPA") or row["TITULO"]).strip(),
                      "media": _existing_media(row.get("MEDIA_PASO"), f"Pasos/{route}/{order}/MEDIA_PASO", optional=True)})
    used_routes = {route for content in contents for route in content["routes"].values()}
    if used_routes != set(orders):
        raise CMSValidationError(f"Rutas/pasos no coinciden: rutas={sorted(used_routes)}, pasos={sorted(orders)}")
    for route, sequence in orders.items():
        if sorted(sequence) != list(range(1, len(sequence) + 1)):
            raise CMSValidationError(f"Orden no consecutivo en ruta {route}: {sorted(sequence)}")

    catalog = []
    media_ids = []
    for row in media_rows:
        media_id = str(row["ID_MEDIO"]).strip()
        media_ids.append(media_id)
        name, category, subcategory = (str(row[key]).strip() for key in ("NOMBRE", "CATEGORIA", "SUBCATEGORIA"))
        catalog.append({"id": media_id, "name": name, "category": category, "subcategory": subcategory,
                        "productImage": _existing_media(row["IMAGEN_PRODUCTO"], f"Medios/{media_id}/IMAGEN_PRODUCTO"),
                        "referenceImage": _existing_media(row["FICHA_REFERENCIA"], f"Medios/{media_id}/FICHA_REFERENCIA"),
                        "search": f"{name} {category} {subcategory} {row.get('FUENTE') or ''}".strip()})
    _unique(media_ids, "Medios/ID_MEDIO")
    if set(media_ids) != content_id_set:
        raise CMSValidationError(
            f"Contenidos/Medios no son 1:1: sin ficha={sorted(content_id_set-set(media_ids))}, "
            f"sin módulo={sorted(set(media_ids)-content_id_set)}"
        )

    campaigns = []
    for row in campaign_rows:
        if not _active(row.get("ACTIVO")):
            continue
        primary, secondary = str(row["ID_PRINCIPAL"]).strip(), str(row["ID_SECUNDARIO"]).strip()
        if primary not in content_ids or secondary not in content_ids:
            raise CMSValidationError(f"Campaña con contenidos desconocidos: {primary}, {secondary}")
        campaigns.append({
            "id": str(row["ID_CAMPANA"]).strip(), "title": str(row["TITULO"]).strip(),
            "subtitle": str(row.get("SUBTITULO") or "").strip(), "start": str(row["FECHA_INICIO"]).strip(),
            "end": str(row["FECHA_FIN"]).strip(), "timezone": str(row["ZONA_HORARIA"]).strip(),
            "primary": primary, "secondary": secondary,
            "resources": [_safe_media(row[key]) for key in ("IMAGEN_CHECKLIST", "IMAGEN_PRACTICAS", "IMAGEN_CONCURSO") if row.get(key)],
        })

    return {"meta": {"version": "5.0.0-core", "catalogItems": len(catalog), "trainingModules": len(contents),
                     "source": "outputs/CMS_Guia_Operativa_CORE.xlsx", "campaigns": campaigns}, "labels": labels, "catalog": catalog,
            "contents": contents, "steps": sorted(steps, key=lambda item: (item["route"], item["order"]))}
