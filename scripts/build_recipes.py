#!/usr/bin/env python3
"""Convierte el CMS Excel en el motor JS consumido por la guía HTML."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def rows_by_header(sheet) -> list[dict[str, object]]:
    headers = [clean(cell.value).upper() for cell in sheet[1]]
    result = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        if any(clean(value) for value in values):
            result.append(row)
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="outputs/CMS_Recetarios_Manuales_Frappuccino.xlsx")
    parser.add_argument("--output", default="data/recipes.js")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    source = root / args.excel
    workbook = load_workbook(source, data_only=False, read_only=True)
    required = {"Recetas", "Variantes", "Pasos", "Tamanos"}
    missing = required - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Faltan pestañas: {', '.join(sorted(missing))}")

    recipe_rows = rows_by_header(workbook["Recetas"])
    variant_rows = rows_by_header(workbook["Variantes"])
    step_rows = rows_by_header(workbook["Pasos"])
    size_rows = rows_by_header(workbook["Tamanos"])

    sizes = [
        {"id": clean(row["ID_TAMANO"]), "label": clean(row["TAMANO"]), "short": clean(row["TAMANO"])[0:1].upper()}
        for row in sorted(size_rows, key=lambda row: int(row["ORDEN"]))
    ]
    size_ids = [size["id"] for size in sizes]
    recipes = []
    for recipe_row in recipe_rows:
        if clean(recipe_row.get("ACTIVA")).upper() != "SI":
            continue
        recipe_id = clean(recipe_row["ID_RECETA"])
        variants = []
        for variant_row in sorted(
            (row for row in variant_rows if clean(row["ID_RECETA"]) == recipe_id and clean(row.get("ACTIVA")).upper() == "SI"),
            key=lambda row: int(row["ORDEN"]),
        ):
            variant_id = clean(variant_row["VARIANTE"])
            matching_steps = sorted(
                (
                    row for row in step_rows
                    if clean(row["ID_RECETA"]) == recipe_id
                    and clean(row["VARIANTE"]) == variant_id
                    and clean(row.get("MOSTRAR")).upper() == "SI"
                ),
                key=lambda row: int(row["ORDEN"]),
            )
            steps = []
            for row in matching_steps:
                steps.append({
                    "order": int(row["ORDEN"]),
                    "icon": clean(row["ICONO"]),
                    "title": clean(row["PASO"]),
                    "detail": clean(row["DETALLE"]),
                    "values": {size_id: clean(row[size_id]) for size_id in size_ids},
                })
            uses_roast = clean(variant_row["USA_ROAST"]).upper() == "SI"
            variants.append({
                "id": variant_id,
                "label": clean(variant_row["ETIQUETA"]),
                "note": "Incluye Frappuccino Roast" if uses_roast else "Sin Roast · empieza con leche",
                "steps": steps,
            })
        recipes.append({
            "id": recipe_id,
            "name": clean(recipe_row["NOMBRE"]),
            "description": clean(recipe_row["DESCRIPCION"]),
            "image": clean(recipe_row["IMAGEN"]),
            "askVariant": clean(recipe_row["PREGUNTA_VARIANTE"]).upper() == "SI",
            "variants": variants,
        })

    if not recipes or any(not recipe["variants"] for recipe in recipes):
        raise ValueError("El CMS no contiene recetas y variantes operativas completas")
    cream = next(variant for recipe in recipes if recipe["id"] == "FRAP_CAJETA" for variant in recipe["variants"] if variant["id"] == "CREAM")
    if not cream["steps"] or "leche" not in cream["steps"][0]["title"].lower() or any("roast" in step["title"].lower() for step in cream["steps"]):
        raise ValueError("Regla inválida: Cajeta Cream debe comenzar con leche y omitir Roast")

    payload = {
        "meta": {"version": "cms-v1", "source": source.name, "recipes": len(recipes)},
        "sizes": sizes,
        "recipes": recipes,
    }
    target = root / args.output
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("window.RECIPE_CMS = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
    print(json.dumps({"status": "ok", "recipes": len(recipes), "variants": sum(len(recipe["variants"]) for recipe in recipes), "steps": sum(len(variant["steps"]) for recipe in recipes for variant in recipe["variants"]), "output": str(target)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
